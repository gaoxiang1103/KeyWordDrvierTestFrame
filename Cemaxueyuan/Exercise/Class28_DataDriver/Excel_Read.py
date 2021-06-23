import openpyxl
from Webkeys import WebKeys

excel = openpyxl.load_workbook('E:\MyProgram\PycharmProjects\Cemaxueyuan\Exercise\Class28\TestCase.xlsx')
sheets = excel.sheetnames
for sheet in sheets:
    sheet_temp = excel[sheet]
    for values in sheet_temp.values:
        if type(values[0]) is int:
            # print(values)
            data = {}
            data['name'] = values[2]
            data['value'] = values[3]
            data['txt'] = values[4]
            print(data)
            for key in list(data.keys()):
                if data[key] is None:
                    del data[key]
            print(data)

            if values[1] == 'open_browser':
                wk = WebKeys(values[4])
            elif 'assert' in values[1]:
                status = getattr(wk, values[1])(**data)
                if status:
                    sheet_temp.cell(row = values[0] + 2, column = 7).value = 'Pass'
                else:
                    sheet_temp.cell(row = values[0] + 2, column = 7).value = 'Falied'
            else:
                getattr(wk,values[1])(**data)

excel.save('E:\MyProgram\PycharmProjects\Cemaxueyuan\Exercise\Class28\TestCase.xlsx')








