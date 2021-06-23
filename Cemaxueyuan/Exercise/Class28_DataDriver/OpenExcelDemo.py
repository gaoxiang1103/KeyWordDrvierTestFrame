"""
读取Excel表格中的Sheet页及单元格的数据
"""
import openpyxl

excel = openpyxl.load_workbook('E:\MyProgram\PycharmProjects\Cemaxueyuan\Exercise\Class28\data.xlsx')
# 定位Sheet页
Sheet = excel['Sheet1']
# 定位单元格
cell = Sheet['A1']
# 读取Sheet页中的所有数据
values = Sheet.values
for v in values:
    print(v)
# 获取所有Sheet页中的单元格数据
sheets = excel.sheetnames
for sheet in sheets:
    print('*' * 20)
    for v in excel[sheet].values:
        print(v)

# 向单元格中写入东西
Sheet.cell(row = 5,column= 5).value = '666'
excel.save('E:\MyProgram\PycharmProjects\Cemaxueyuan\Exercise\Class28\data.xlsx')


